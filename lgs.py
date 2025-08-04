from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import time
import pandas as pd
import openpyxl
from io import StringIO
import pandas as pd
from PyQt5.QtWidgets import QApplication, QFileDialog, QMessageBox, QLabel, QLineEdit, QVBoxLayout, QPushButton, QDialog
import sys
from selenium.webdriver.support.ui import Select
import logging
import os

chrome_options = Options()
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(options=chrome_options)

logging.basicConfig(filename='error_log.txt', level=logging.ERROR)

def save_progress(index):
    with open('progress.txt', 'w') as f:
        f.write(str(index))

def load_progress():
    try:
        with open('progress.txt', 'r') as f:
            return int(f.read())
    except FileNotFoundError:
        return 0

def delete_progress():
    try:
        os.remove('progress.txt')
    except FileNotFoundError:
        pass

def dosya_ac():
    global write_file, excel_file, read_sheet_name, write_sheet_name
    ogrenci_dosyası_secme_mesajı()
    options = QFileDialog.Options()
    excel_file, _ = QFileDialog.getOpenFileName(None, "Dosya Seç", "", "All Files (*)", options=options)
    read_workbook = openpyxl.load_workbook(excel_file)
    read_sheet_name = read_workbook.sheetnames[0]
    sonuc_dosyası_secme_mesajı()
    write_file, _ = QFileDialog.getOpenFileName(None, "Dosya Seç", "", "All Files (*)", options=options)
    write_workbook = openpyxl.load_workbook(write_file)
    write_sheet_name = write_workbook.sheetnames[0]

def tercih_yapmayan_ogrenci_veri():
    birthday = str(bDay) + "." + str(bMonth) + "." + str(bYear)
    data = {
        '0': ["T.C. Kimlik No", "Adı Soyadı", "Doğum Tarihi", "Sonuç"],
        '1': [str(id), name, birthday, 'T.C. Kimlik Numaranızı/Doğum Tarihinizi Yanlış Girdiniz veya Tercih Başvurunuz Bulunmamaktadır!']
    }
    data_frame = pd.DataFrame(data)
    tablo_verileri.append(data_frame)

def tablo_verilerini_al():
    """
    Sayfadaki TÜM tabloları almak için geliştirilmiş fonksiyon
    """
    try:
        print("Tablo verilerini alınıyor...")
        page_tables_found = 0
        
        # Sayfadaki TÜM tabloları bul ve al
        try:
            # Önce tüm table elementlerini bul
            all_tables = driver.find_elements(By.TAG_NAME, "table")
            print(f"Sayfada {len(all_tables)} adet table elementi bulundu")
            
            for i, table in enumerate(all_tables):
                try:
                    # Tablonun görünür ve içi dolu olup olmadığını kontrol et
                    if table.is_displayed():
                        table_html = table.get_attribute('outerHTML')
                        if table_html and len(table_html.strip()) > 100:  # Minimum içerik kontrolü
                            print(f"Tablo {i+1} işleniyor...")
                            
                            # Pandas ile parse et
                            table_dataframes = pd.read_html(StringIO(table_html))
                            
                            for j, df in enumerate(table_dataframes):
                                if not df.empty and df.shape[0] > 0:
                                    print(f"Tablo {i+1}-{j+1} boyutu: {df.shape}")
                                    print(f"İlk birkaç satır:")
                                    print(df.head(3))
                                    tablo_verileri.append(df)
                                    page_tables_found += 1
                                    
                except Exception as table_error:
                    print(f"Tablo {i+1} parse hatası: {table_error}")
                    continue
            
        except Exception as e:
            print(f"Table elementleri bulma hatası: {e}")
        
        # Eğer hiç tablo bulunamadıysa, sayfadaki diğer yapıları kontrol et
        if page_tables_found == 0:
            print("HTML tablosu bulunamadı, div/span tabanlı yapıları arıyor...")
            
            # Div tabanlı tablo yapılarını ara
            div_tables = driver.find_elements(By.XPATH, "//div[contains(@class, 'table') or contains(@class, 'result') or contains(@class, 'sonuc')]")
            
            for div_table in div_tables:
                try:
                    if div_table.is_displayed():
                        div_html = div_table.get_attribute('outerHTML')
                        if "T.C. Kimlik" in div_html or "Adı Soyadı" in div_html:
                            print("Div tabanlı tablo bulundu, manuel parse ediliyor...")
                            # Manuel veri çıkarımı yap
                            manual_data = extract_manual_data_from_page()
                            if manual_data:
                                tablo_verileri.append(manual_data)
                                page_tables_found += 1
                            break
                except Exception as div_error:
                    print(f"Div tablo hatası: {div_error}")
                    continue
        
        # Son çare: Manuel veri çıkarımı
        if page_tables_found == 0:
            print("Hiçbir tablo bulunamadı, manuel veri çıkarımı yapılıyor...")
            manual_data = extract_manual_data_from_page()
            if manual_data is not None:
                tablo_verileri.append(manual_data)
                page_tables_found += 1
        
        print(f"Toplam {page_tables_found} adet tablo verisi alındı")
        
        if page_tables_found == 0:
            print("UYARI: Hiçbir tablo verisi alınamadı!")
            # Debug dosyaları kaydet
            try:
                driver.save_screenshot('tablo_bulunamadi.png')
                with open('sayfa_kaynagi.html', 'w', encoding='utf-8') as f:
                    f.write(driver.page_source)
                print("Debug dosyaları kaydedildi: tablo_bulunamadi.png, sayfa_kaynagi.html")
            except:
                pass
        
        return pd.concat(tablo_verileri).drop_duplicates() if tablo_verileri else pd.DataFrame()
        
    except Exception as e:
        print(f"Tablo alma genel hatası: {e}")
        return pd.DataFrame()

def extract_manual_data_from_page():
    """
    Sayfadan manuel veri çıkarımı yapar
    """
    try:
        page_source = driver.page_source
        birthday = str(bDay) + "." + str(bMonth) + "." + str(bYear)
        
        # Farklı sonuç metinlerini ara
        sonuc_text = "Veri alınamadı"
        
        if "Kayıt Alınmızda Yer Alan" in page_source:
            try:
                sonuc_element = driver.find_element(By.XPATH, "//*[contains(text(), 'Kayıt Alınmızda Yer Alan')]")
                sonuc_text = sonuc_element.text
            except:
                sonuc_text = "Kayıt Alınmızda Yer Alan 3 Nolu Tercihinize Yerleştiriliz"
        elif "BURSA" in page_source and "Meslek" in page_source:
            try:
                # Yerleştirilen okul bilgisini bul
                okul_elements = driver.find_elements(By.XPATH, "//*[contains(text(), 'BURSA') or contains(text(), 'Meslek') or contains(text(), 'Lisesi')]")
                if okul_elements:
                    sonuc_text = okul_elements[0].text
            except:
                sonuc_text = "Meslek lisesine yerleştirildi"
        elif "Yerleştirme" in page_source:
            try:
                sonuc_element = driver.find_element(By.XPATH, "//*[contains(text(), 'Yerleştirme')]")
                sonuc_text = sonuc_element.text
            except:
                sonuc_text = "Yerleştirme yapıldı"
        
        # Manuel dataframe oluştur
        data = {
            'T.C. Kimlik No': [str(id)],
            'Adı Soyadı': [name],
            'Doğum Tarihi': [birthday],
            'Sonuç': [sonuc_text]
        }
        
        return pd.DataFrame(data)
        
    except Exception as e:
        print(f"Manuel veri çıkarım hatası: {e}")
        return None

def tablo_verilerini_yaz():
    if not tablo_verileri:
        print("Yazılacak tablo verisi yok!")
        return
        
    concatenated_df = pd.concat(tablo_verileri)
    write_dolu_satir_sayisi = read_excel_row_count(write_file, write_sheet_name)
    # Eğer write_dolu_satir_sayisi 0 ise, startrow 0 olur, aksi takdirde startrow write_dolu_satir_sayisi olur
    startrow = write_dolu_satir_sayisi if write_dolu_satir_sayisi == 0 else write_dolu_satir_sayisi + 1
    print(f"Excel'e yazılıyor, startrow: {startrow}")
    
    with pd.ExcelWriter(write_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        concatenated_df.to_excel(writer, sheet_name=write_sheet_name, index=False, startrow=startrow, header=write_dolu_satir_sayisi==0)

    # Clear the list after writing to the file
    tablo_verileri.clear()    

def yeni_sorgu():
    """
    Yeni sorgu için sayfayı yenileme ve input alanlarını temizleme - geliştirilmiş versiyon
    """
    try:
        print("Yeni sorgu için sayfa yenileniyor...")
        
        # Önce "Yeni Sorgu" linkini aramaya çalış
        yeni_sorgu_selectors = [
            (By.PARTIAL_LINK_TEXT, "Yeni Sorgu"),
            (By.LINK_TEXT, "Yeni Sorgu"),
            (By.XPATH, "//a[contains(text(), 'Yeni Sorgu')]"),
            (By.XPATH, "//a[contains(text(), 'yeni sorgu')]"),
            (By.XPATH, "//input[@value='Yeni Sorgu']"),
            (By.XPATH, "//button[contains(text(), 'Yeni Sorgu')]")
        ]
        
        link_found = False
        for selector_type, selector_value in yeni_sorgu_selectors:
            try:
                yeni_sorgu_element = driver.find_element(selector_type, selector_value)
                if yeni_sorgu_element.is_displayed():
                    yeni_sorgu_element.click()
                    print("Yeni Sorgu linkine tıklandı")
                    link_found = True
                    time.sleep(2)
                    break
            except NoSuchElementException:
                continue
        
        # Eğer link bulunamazsa, browser back kullan
        if not link_found:
            print("Yeni Sorgu linki bulunamadı, geri gidiliyor...")
            driver.back()
            time.sleep(2)
            
            # Geri gittikten sonra da çalışmazsa sayfayı yenile
            if "sonuc" not in driver.current_url.lower():
                print("Ana sayfaya geri dönülemedi, yenileniyor...")
                driver.refresh()
                time.sleep(3)
        
        # ÖNEMLI: Input alanlarını temizle
        print("Input alanları temizleniyor...")
        temizle_input_alanlari()
        
        print(f"Mevcut URL: {driver.current_url}")
        
    except Exception as e:
        print(f"Yeni sorgu hatası: {e}")
        # En son çare olarak sayfayı yenile
        try:
            driver.refresh()
            time.sleep(3)
            temizle_input_alanlari()
        except:
            pass

def temizle_input_alanlari():
    """
    Sayfadaki tüm input alanlarını temizler
    """
    try:
        print("Input alanları temizleniyor...")
        
        # Tüm text input alanlarını bul ve temizle
        text_inputs = driver.find_elements(By.XPATH, "//input[@type='text' or not(@type)]")
        
        for inp in text_inputs:
            try:
                if inp.is_displayed() and inp.is_enabled():
                    inp.clear()
                    print(f"Input temizlendi: {inp.get_attribute('name') or inp.get_attribute('id') or 'Bilinmeyen'}")
            except Exception as inp_error:
                print(f"Input temizleme hatası: {inp_error}")
                continue
        
        # Dropdown'ları da sıfırla (ilk seçeneği seç)
        dropdowns = driver.find_elements(By.TAG_NAME, "select")
        
        for dropdown in dropdowns:
            try:
                if dropdown.is_displayed() and dropdown.is_enabled():
                    select_obj = Select(dropdown)
                    # İlk seçeneği seç (genelde boş veya varsayılan)
                    if len(select_obj.options) > 0:
                        select_obj.select_by_index(0)
                        dropdown_name = dropdown.get_attribute('name') or dropdown.get_attribute('id') or 'Bilinmeyen'
                        print(f"Dropdown sıfırlandı: {dropdown_name}")
                        
                        # Yıl dropdown'u için özel kontrol
                        if dropdown_name in ['YlL', 'YIL', 'yil', 'yl']:
                            print(f"YIL dropdown bulundu: name='{dropdown_name}'")
                            
            except Exception as dropdown_error:
                print(f"Dropdown sıfırlama hatası: {dropdown_error}")
                continue
        
        # Güvenlik kodu input'unu da temizle
        try:
            guvenlik_inputs = driver.find_elements(By.XPATH, "//input[contains(@id, 'GUVENLIK') or contains(@id, 'gkodu') or contains(@name, 'guvenlik')]")
            for guv_inp in guvenlik_inputs:
                if guv_inp.is_displayed() and guv_inp.is_enabled():
                    guv_inp.clear()
                    print("Güvenlik kodu input'u temizlendi")
        except:
            pass
        
        print("Input alanları temizleme tamamlandı")
        
    except Exception as e:
        print(f"Input temizleme genel hatası: {e}")

def find_input_by_attribute(driver, attribute, value):
    return driver.find_element(By.CSS_SELECTOR, f'input[{attribute}="{value}"]')

def find_element_with_fallbacks(driver, selectors_list, element_name):
    """
    Birden fazla selector deneyerek elementi bulmaya çalışır
    """
    for selector_type, selector_value in selectors_list:
        try:
            element = driver.find_element(selector_type, selector_value)
            print(f"{element_name} bulundu: {selector_type} = '{selector_value}'")
            return element
        except NoSuchElementException:
            continue
    
    # Hiçbiri bulunamazsa hata fırlat
    print(f"HATA: {element_name} bulunamadı. Denenen selectors:")
    for selector_type, selector_value in selectors_list:
        print(f"  - {selector_type}: '{selector_value}'")
    
    # Sayfanın kaynak kodunu kontrol etmek için
    print("Sayfa kaynağında TC kimlik ile ilgili inputlar:")
    page_source = driver.page_source.lower()
    if 'kimlik' in page_source:
        print("✓ 'kimlik' kelimesi sayfada bulundu")
    if 'tc' in page_source:
        print("✓ 'tc' kelimesi sayfada bulundu")
    
    raise NoSuchElementException(f"{element_name} için hiçbir selector çalışmadı")

def fill_date_dropdowns():
    """
    Tarih dropdown'larını güvenli şekilde doldur
    """
    dropdown_selectors = [
        ("GUN", str(bDay), "Gün"),
        ("AY", str(bMonth), "Ay"), 
        ("YIL", str(bYear), "Yıl")
    ]
    
    for name_value, select_value, field_name in dropdown_selectors:
        try:
            # Yıl için özel durumlar (YIL, YlL, yil, yl gibi varyasyonlar)
            if name_value == "YIL":
                possible_names = ["YIL", "YlL", "yil", "yl", "YEAR", "year"]
            else:
                possible_names = [name_value, name_value.lower(), name_value.upper()]
            
            dropdown_found = False
            
            for possible_name in possible_names:
                try:
                    dropdown = Select(driver.find_element(By.NAME, possible_name))
                    dropdown.select_by_value(select_value)
                    print(f"{field_name} seçildi: {select_value} (name='{possible_name}')")
                    dropdown_found = True
                    break
                except NoSuchElementException:
                    continue
            
            if not dropdown_found:
                # ID ile de dene
                try:
                    dropdown = Select(driver.find_element(By.ID, name_value))
                    dropdown.select_by_value(select_value)
                    print(f"{field_name} (ID ile) seçildi: {select_value}")
                    dropdown_found = True
                except NoSuchElementException:
                    pass
            
            # Son çare: XPath ile ara
            if not dropdown_found:
                try:
                    xpath_selectors = [
                        f"//select[contains(@name, '{field_name.lower()}')]",
                        f"//select[contains(@id, '{field_name.lower()}')]",
                        "//select[position()=3]" if field_name == "Yıl" else None
                    ]
                    
                    for xpath in xpath_selectors:
                        if xpath:
                            try:
                                dropdown = Select(driver.find_element(By.XPATH, xpath))
                                dropdown.select_by_value(select_value)
                                print(f"{field_name} (XPath ile) seçildi: {select_value}")
                                dropdown_found = True
                                break
                            except:
                                continue
                except:
                    pass
            
            if not dropdown_found:
                print(f"UYARI: {field_name} dropdown'u bulunamadı")
                # Debug için mevcut dropdown'ları listele
                try:
                    all_selects = driver.find_elements(By.TAG_NAME, "select")
                    print(f"Sayfadaki dropdown'lar:")
                    for i, sel in enumerate(all_selects):
                        sel_name = sel.get_attribute("name") or "N/A"
                        sel_id = sel.get_attribute("id") or "N/A"
                        print(f"  {i+1}. name='{sel_name}', id='{sel_id}'")
                except:
                    pass
                    
        except Exception as e:
            print(f"UYARI: {field_name} seçilirken hata: {e}")

def debug_page_elements():
    """
    Sayfadaki tüm input ve select elementlerini listele
    """
    print("\n=== SAYFA ELEMENT ANALIZI ===")
    
    # Tüm input elementleri
    inputs = driver.find_elements(By.TAG_NAME, "input")
    print(f"Toplam input sayısı: {len(inputs)}")
    
    for i, inp in enumerate(inputs):
        try:
            inp_type = inp.get_attribute("type") or "text"
            inp_id = inp.get_attribute("id") or "N/A"
            inp_name = inp.get_attribute("name") or "N/A"
            inp_placeholder = inp.get_attribute("placeholder") or "N/A"
            inp_class = inp.get_attribute("class") or "N/A"
            
            print(f"Input {i+1}: type={inp_type}, id={inp_id}, name={inp_name}, placeholder={inp_placeholder}, class={inp_class}")
        except:
            print(f"Input {i+1}: Element analiz edilemedi")
    
    # Tüm select elementleri
    selects = driver.find_elements(By.TAG_NAME, "select")
    print(f"\nToplam select sayısı: {len(selects)}")
    
    for i, sel in enumerate(selects):
        try:
            sel_id = sel.get_attribute("id") or "N/A"
            sel_name = sel.get_attribute("name") or "N/A"
            sel_class = sel.get_attribute("class") or "N/A"
            
            print(f"Select {i+1}: id={sel_id}, name={sel_name}, class={sel_class}")
        except:
            print(f"Select {i+1}: Element analiz edilemedi")
    
    print("=== ANALIZ TAMAMLANDI ===\n")

def giris_elemanları():
    try:
        # TC Kimlik No için farklı selector alternatifleri
        tc_selectors = [
            (By.XPATH, '//input[contains(@placeholder, "T.C. Kimlik")]'),
            (By.XPATH, '//input[contains(@placeholder, "Kimlik")]'),
            (By.XPATH, '//input[contains(@placeholder, "T.C.")]'),
            (By.XPATH, '//input[@id="TC_KIMLIK_NO"]'),
            (By.XPATH, '//input[@id="TCNO"]'),
            (By.XPATH, '//input[@id="ADAY_NO"]'),
            (By.XPATH, '//input[@name="TC_KIMLIK_NO"]'),
            (By.XPATH, '//input[@name="TCNO"]'),
            (By.XPATH, '//input[contains(@class, "tc")]'),
            (By.CSS_SELECTOR, 'input[placeholder*="Kimlik"]'),
            (By.CSS_SELECTOR, 'input[placeholder*="T.C"]'),
        ]
        
        # Okul No için farklı selector alternatifleri
        okul_selectors = [
            (By.XPATH, '//input[contains(@placeholder, "Okul")]'),
            (By.XPATH, '//input[contains(@placeholder, "OKUL")]'),
            (By.XPATH, '//input[@id="OKULNO"]'),
            (By.XPATH, '//input[@id="GUVENLIKNUMARASI"]'),
            (By.XPATH, '//input[@name="OKULNO"]'),
            (By.XPATH, '//input[@name="OKUL_NO"]'),
            (By.XPATH, '//input[contains(@class, "okul")]'),
            (By.CSS_SELECTOR, 'input[placeholder*="Okul"]'),
            (By.CSS_SELECTOR, 'input[placeholder*="OKUL"]'),
        ]
        
        # Elementleri bul
        tcInput = find_element_with_fallbacks(driver, tc_selectors, "TC Kimlik Input")
        okulNoInput = find_element_with_fallbacks(driver, okul_selectors, "Okul No Input")
        
        # Verileri temizle ve gir
        tcInput.clear()
        tcInput.send_keys(str(id))
        print(f"TC Kimlik girildi: {id}")
        
        okulNoInput.clear()
        okulNoInput.send_keys(str(okulNo))
        print(f"Okul No girildi: {okulNo}")
        
        # Dropdown'lar için güvenli seçim
        fill_date_dropdowns()
        
    except Exception as e:
        print(f"HATA - Giriş elemanları doldurulurken: {e}")
        print(f"Mevcut URL: {driver.current_url}")
        print(f"Sayfa başlığı: {driver.title}")
        
        # Hata durumunda sayfanın HTML'ini dosyaya kaydet
        with open('hata_sayfa.html', 'w', encoding='utf-8') as f:
            f.write(driver.page_source)
        print("Sayfa kaynağı 'hata_sayfa.html' dosyasına kaydedildi")
        
        raise

def uyarı_mesaj_guvenlik_metin():
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Critical)
    msg.setText("Girdiğiniz güvenlik kodunda bir hata oluştu! sadece güvenlik kodunu elle giriniz ve sonra uyarı mesajındaki tamama basınız")
    msg.setWindowTitle("Hata Mesajı")
    msg.exec_()

def bilgi_mesajı():
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setText("Güvenlik kodunu giriniz ve sonra uyarı mesajındaki tamama basınız")
    msg.setWindowTitle("Dikkat")
    msg.exec_()

def sınava_girmeyen_ogrenci_mesajı():
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setText(f"{name} adlı öğrenci sınava girmemiştir veya tercih başvurusunda bulunmamıştır. Tamam basarak sistemin işlemesine devam ediniz")
    msg.setWindowTitle("Dikkat")
    msg.exec_()

def bitis_mesajı():
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setText("Öğrencilerin sonuçlarını alma işlemi tamamlanmıştır.")
    msg.setWindowTitle("Dikkat")
    msg.exec_()

def ogrenci_dosyası_secme_mesajı():
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setText("Lgs sonuçları çekilecek olan öğrencilerin excel dosyasını açılan ekrandan seçiniz")
    msg.setWindowTitle("Dikkat")
    msg.exec_()

def sonuc_dosyası_secme_mesajı():
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setText("Lgs sonuçlarının kaydedileceği excel dosyasını açılan ekrandan seçiniz")
    msg.setWindowTitle("Dikkat")
    msg.exec_()

def is_element_present(driver, by, selector):
    try:
        driver.find_element(by, selector)
        return True
    except NoSuchElementException:
        return False

def read_excel_row(file_path, sheet_name, row_index):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    row_data = df.iloc[row_index].values.tolist()
    return row_data

def read_excel_row_count(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    dolu_satir_sayisi = df.shape[0]
    return dolu_satir_sayisi

def başarılı_sorgu_sonuc():
    """
    Başarılı sorgu sonucunda tablo verilerini al ve yeni sorguya geç
    """
    print("Başarılı sorgu sonucu işleniyor...")
    
    # Sayfanın yüklenmesini bekle
    time.sleep(2)
    
    # Tablo verilerini al
    tablo_verilerini_al()
    
    # Yeni sorgu için geri dön
    yeni_sorgu()

def get_input():
    dialog = QDialog()
    dialog.setWindowTitle('LGS SONUÇ URL')

    label = QLabel('Sonuçların alınacağı url giriniz:')
    input_field = QLineEdit()

    def on_ok_clicked():
        global input_value
        input_value = input_field.text()
        dialog.accept()

    ok_button = QPushButton('Tamam')
    ok_button.clicked.connect(on_ok_clicked)

    layout = QVBoxLayout()
    layout.addWidget(label)
    layout.addWidget(input_field)
    layout.addWidget(ok_button)

    dialog.setLayout(layout)
    dialog.setFixedSize(300, 150)

    dialog.exec_()
    return input_value

def sonucları_al():
    global name, bDay, bMonth, bYear, id, okulNo
    start_index = load_progress()
    
    for i in range(start_index, dolu_satir_sayisi):
        try:
            row_index = i  
            row_data = read_excel_row(excel_file, read_sheet_name, row_index)
            name, bDay, bMonth, bYear, id, okulNo = row_data[:6]
            
            print(f"\n--- İşleniyor: {name} ({i+1}/{dolu_satir_sayisi}) ---")
            
            if int(bDay) < 10:
                bDay = "0" + str(bDay)
            if int(bMonth) < 10:
                bMonth = "0" + str(bMonth)
            
            try:
                giris_elemanları()
            except NoSuchElementException:
                print("Element bulunamadı - Debug yapılıyor...")
                debug_page_elements()
                raise
                
            guvenlık_kodu_selector = (By.XPATH, '//input[(@id="GUVENLIKKODU") or (@id="gkodu") or (@name="GUVENLIKKODU") or contains(@placeholder, "Güvenlik")]')
            hata_kodu_selector = (By.ID, "hata")
            yeni_sayfada_hata_kodu_selector = (By.XPATH, "//p[@align='center']")
            
            # Güvenlik kodu var mı kontrol et
            print("Güvenlik kodu aranıyor...")
            guvenlik_kodu_var = is_element_present(driver, *guvenlık_kodu_selector)
            print(f"Güvenlik kodu bulundu: {guvenlik_kodu_var}")
            
            if guvenlik_kodu_var:
                print("Güvenlik kodu var - kullanıcı girmeli")
                bilgi_mesajı()  # Kullanıcıya güvenlik kodunu girmesi için uyarı ver ve bekle
                print("Kullanıcı güvenlik kodunu girdi, devam ediliyor...")
            
            # Tamam butonuna bas
            tamam_button = driver.find_element(By.NAME, "Submit")
            tamam_button.click()
            time.sleep(3)
            
            if guvenlik_kodu_var:
                
                if is_element_present(driver, *hata_kodu_selector):
                    hatakodutext = driver.find_element(By.XPATH, "//*[@id='hata']")
                    if hatakodutext.text == "Güvenlik Kodunu yanlış girdiniz!":
                        uyarı_mesaj_guvenlik_metin()
                        try:
                            giris_elemanları()
                        except NoSuchElementException:
                            print("Element bulunamadı - Debug yapılıyor...")
                            debug_page_elements()
                            raise
                        time.sleep(1)
                        tamam_button = driver.find_element(By.NAME, "Submit")
                        tamam_button.click()
                        time.sleep(3)
                        if is_element_present(driver, *hata_kodu_selector):
                            sınava_girmeyen_ogrenci_mesajı()
                            tercih_yapmayan_ogrenci_veri()
                            yeni_sorgu()
                            continue
                    else:
                        sınava_girmeyen_ogrenci_mesajı()
                        tercih_yapmayan_ogrenci_veri()
                        yeni_sorgu()
                        continue
                
                if is_element_present(driver, *yeni_sayfada_hata_kodu_selector):
                    yeni_hatakodutext = driver.find_element(By.XPATH, '//p[1]')
                    if yeni_hatakodutext.text in ["T.C. Kimlik Numaranızı/Doğum Tarihinizi Yanlış Girdiniz veya Tercih Başvurunuz Bulunmamaktadır!", "T.C. Kimlik Numaranızı veya Doğum Tarihinizi Yanlış Girdiniz!"]:
                        sınava_girmeyen_ogrenci_mesajı()
                        tercih_yapmayan_ogrenci_veri()
                        yeni_sorgu()
                        continue
                    else:
                        yeni_sorgu()
                        try:
                            giris_elemanları()
                        except NoSuchElementException:
                            print("Element bulunamadı - Debug yapılıyor...")
                            debug_page_elements()
                            raise
                        uyarı_mesaj_guvenlik_metin()
                        tamam_button = driver.find_element(By.NAME, "Submit")
                        tamam_button.click()
                        time.sleep(3)
                        if is_element_present(driver, *yeni_sayfada_hata_kodu_selector):
                            sınava_girmeyen_ogrenci_mesajı()
                            tercih_yapmayan_ogrenci_veri()
                            yeni_sorgu()
                            continue
                
                # Başarılı durum - tablo verilerini al
                başarılı_sorgu_sonuc()
            
            else:
                print("Güvenlik kodu yok - direkt gönderiliyor")
                tamam_button = driver.find_element(By.NAME, "Submit")
                tamam_button.click()
                time.sleep(3)  # Daha uzun bekleme süresi
                
                if is_element_present(driver, *hata_kodu_selector):
                    sınava_girmeyen_ogrenci_mesajı()
                    tercih_yapmayan_ogrenci_veri()
                    yeni_sorgu()
                    continue
                    
                if is_element_present(driver, *yeni_sayfada_hata_kodu_selector):
                    sınava_girmeyen_ogrenci_mesajı()
                    tercih_yapmayan_ogrenci_veri()
                    yeni_sorgu()
                    continue
                    
                # Başarılı durum - tablo verilerini al
                başarılı_sorgu_sonuc()
            
            save_progress(i + 1)  # İlerlemeyi her adımda kaydet
            
            if (i + 1) % 5 == 0:
                tablo_verilerini_yaz()
                
        except Exception as e:
            logging.error(f"Hata oluştu (öğrenci indeksi {i}): {str(e)}")
            save_progress(i)  # Hata durumunda son başarılı konumu kaydet
            tablo_verilerini_yaz()  # Hata durumunda mevcut verileri kaydet
            raise  # Hatayı yeniden fırlat
    
    tablo_verilerini_yaz()  # Son kez kaydet

   
if __name__ == "__main__":
    app = QApplication(sys.argv)
    get_input()
    driver = webdriver.Chrome(options=chrome_options)  # options ekleyin
    driver.get(input_value)
    #driver.maximize_window()
    driver.implicitly_wait(1)
    dosya_ac()
    dolu_satir_sayisi = read_excel_row_count(excel_file, read_sheet_name)
    tablo_verileri = []
    try:
        
        sonucları_al()
        delete_progress()
    except Exception as e:
        logging.error(f"Hata oluştu: {str(e)}")
    bitis_mesajı()
    driver.quit()
    app.quit()
    sys.exit(app.exec_())
